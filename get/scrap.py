from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver import ActionChains
import time
from selenium.webdriver.common.actions.wheel_input import ScrollOrigin
import pandas as pd
from selenium.webdriver.support import expected_conditions as EC
import os
import openpyxl
import shutil


class Scrap(webdriver.Chrome):

    def __init__(self, driver_path=r'C:\selenium-driver', teardown=False):
        options = Options()
        experimental_options = {"detach": True, "excludeSwitches": ["enable-logging"]}
        options.add_experimental_option("prefs", experimental_options)
        self.driver_path = driver_path
        self.teardown = teardown
        super().__init__(options=options)
        self.implicitly_wait(10)

    def __exit__(self, exc_type, exc_val, exc_tb):
        if self.teardown:
            self.quit()

    def land_first_page(self):
        self.get('https://www.tradingview.com/')
        self.maximize_window()


    def table(self):
        try:
            table = self.find_element(
                By.XPATH, '//*[@id="bottom-area"]/div[4]/div/div[2]/div/div/table')
            tbody = table.find_element(By.TAG_NAME, 'tbody')
            tr = tbody.find_elements(By.TAG_NAME, 'tr')
            trade = []
            entry_type = []
            exit_type = []
            entry_signal = []
            exit_signal = []
            entry_date_time = []
            exit_date_time = []
            entry_price = []
            exit_price = []
            contracts = []
            profit = []
            cum_profit = []
            run_up = []
            drawdown = []
            i = 1
            while i <= 9:
                try:
                    trade.append(tr[i].find_element(By.XPATH, './td[1]/span').text)
                    entry_type.append(tr[i].find_element(
                        By.XPATH, './td[2]/div[1]/div/span').text)
                    exit_type.append(tr[i].find_element(
                        By.XPATH, './td[2]/div[2]/div/span').text)
                    entry_signal.append(tr[i].find_element(
                        By.XPATH, './td[3]/div[1]/div/span').text)
                    exit_signal.append(tr[i].find_element(
                        By.XPATH, './td[3]/div[2]/div/span').text)
                    entry_date_time.append(tr[i].find_element(
                        By.XPATH, './td[4]/div[1]/div/span').text)
                    exit_date_time.append(tr[i].find_element(
                        By.XPATH, './td[4]/div[2]/div/span').text)
                    entry_price.append(tr[i].find_element(
                        By.XPATH, './td[5]/div[1]/div/span').text)
                    exit_price.append(tr[i].find_element(
                        By.XPATH, './td[5]/div[2]/div/span').text)
                    contracts.append(tr[i].find_element(
                        By.XPATH, './td[6]/div/span').text)
                    profit.append(tr[i].find_element(
                        By.XPATH, './td[7]/div/div[1]').text)
                    cum_profit.append(tr[i].find_element(
                        By.XPATH, './td[8]/div/div[1]').text)
                    run_up.append(tr[i].find_element(
                        By.XPATH, './td[9]/div/div[1]').text)
                    drawdown.append(tr[i].find_element(
                        By.XPATH, './td[10]/div/div[1]').text)
                except:
                    pass
                i += 1
            print(trade)
            return [trade, entry_type, exit_type, entry_signal, exit_signal, entry_date_time, exit_date_time, entry_price, exit_price, contracts, profit, cum_profit]
        except:
            print('Table Not Found')


    def scroll(self):
        table = self.find_element(
            By.XPATH, '//*[@id="bottom-area"]/div[4]/div/div[2]/div/div/table')
        scroll_origin = ScrollOrigin.from_element(table)
        ActionChains(self).scroll_from_origin(scroll_origin, 0, 882).perform()
  

    def save_to_excel(self,data,file_name):
        shutil.copy2('Trading view Trade Log.xlsx', f'{file_name}.xlsx')

        workbook = openpyxl.load_workbook(f'{file_name}.xlsx')
        sheet = workbook.active

        for i, trade in enumerate(data['Trade']):
            row = sheet.max_row + 1
            sheet.cell(row=row, column=2, value=trade)
            sheet.cell(row=row, column=3, value=data['Entry Signal'][i])
            sheet.cell(row=row, column=4, value=data['Entry Date'][i])
            sheet.cell(row=row, column=5, value=data['Entry Price'][i])
            sheet.cell(row=row, column=6, value=data['Exit signal'][i])
            sheet.cell(row=row, column=7, value=data['Exit Date'][i])
            sheet.cell(row=row, column=8, value=data['Exit Price'][i])
            sheet.cell(row=row, column=9, value=data['Profit'][i])
        
        workbook.save(f'{file_name}.xlsx')



    def collect_data(self, file_name):
        try: 
            i = 0
            data = {}
            trade = []
            entry_type = []
            exit_type = []
            entry_signal = []
            exit_signal = []
            entry_date_time = []
            exit_date_time = []
            entry_price = []
            exit_price = []
            contracts = []
            profit = []
            cum_profit = []
            table = self.find_element(
            By.XPATH, '//*[@id="bottom-area"]/div[4]/div/div[2]/div/div/table')
            tr = table.find_elements(By.TAG_NAME, 'tr')
            h = tr[-1].size['height']
            while h > -882:
                
                table_data = self.table()
                if i > 0:
                    if set(table_data[0]).intersection(trade):
                        dup_idx = trade.index(table_data[0][0])  # find index of first duplicate
                        trade = trade[:dup_idx]  # remove items from where duplicates start
                        entry_type = entry_type[:dup_idx]
                        exit_type = exit_type[:dup_idx]
                        entry_signal = entry_signal[:dup_idx]
                        exit_signal = exit_signal[:dup_idx]
                        entry_date_time = entry_date_time[:dup_idx]
                        exit_date_time = exit_date_time[:dup_idx]
                        entry_price = entry_price[:dup_idx]
                        exit_price = exit_price[:dup_idx]
                        contracts = contracts[:dup_idx]
                        profit = profit[:dup_idx]
                        cum_profit = cum_profit[:dup_idx]

                trade += table_data[0]
                entry_type += table_data[1]
                exit_type += table_data[2]
                entry_signal += table_data[3]
                exit_signal += table_data[4]
                entry_date_time += table_data[5]
                exit_date_time += table_data[6]
                entry_price += table_data[7]
                exit_price += table_data[8]
                contracts += table_data[9]
                profit += table_data[10]
                cum_profit += table_data[11]
                self.scroll()
                time.sleep(2.5)
                i += 1
                h -= 882
            data['Trade'] = trade[::-1]
            data['Entry Type'] = entry_type[::-1]
            data['Exit Type'] = exit_type[::-1]
            data['Entry Signal'] = entry_signal[::-1]
            data['Exit signal'] = exit_signal[::-1]
            data['Entry Date'] = entry_date_time[::-1]
            data['Exit Date'] = exit_date_time[::-1]
            data['Entry Price'] = entry_price[::-1]
            data['Exit Price'] = exit_price[::-1]
            data['Contracts'] = contracts[::-1]
            data['Profit'] = profit[::-1]
        
            self.save_to_excel(data,file_name)
        except:
            print('Error while fetching data. Make sure there is a table in your page')
